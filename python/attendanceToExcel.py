import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from datetime import datetime
import sys
import os

# --------- 인자 전달 받기 ----------
csv_path = sys.argv[1]               # attendance.csv 경로
template_path = sys.argv[2]          # 템플릿 엑셀 경로
output_dir = sys.argv[3]             # 출력 폴더
month_str = sys.argv[4]              # 예: "2025-05"

# 날짜 정보 파싱
year, month = map(int, month_str.split('-'))
target_month_display = f"{year}년 {month:02d}월"
target_month_code = f"{year}{month:02d}"

# --------- 파일 로딩 ----------
df = pd.read_csv(csv_path)
df['date'] = pd.to_datetime(df['date'])
df = df[df['date'].dt.strftime('%Y-%m') == month_str]


# 선택한 월 필터링
target_month = pd.to_datetime(month_str)
df = df[df['date'].dt.to_period('M') == target_month.to_period('M')]

# 템플릿 불러오기 
wb = openpyxl.load_workbook(template_path)
ws = wb.active

# A1 제목 설정
ws['A1'] = f"■ {target_month_display} 근무 현황 ■"

# 날짜 → 열 변환 (E~AI)
date_to_col = {day: get_column_letter(4 + day) for day in range(1, 32)}

# 이름 → 시작 행
name_to_row = {}
row = 5
while True:
    cell = ws[f'B{row}']
    if cell.value is None:
        break
    name_to_row[cell.value.strip()] = row
    row += 4

# 숫자로 입력되게 변환하는 함수
def set_number_or_blank(cell, value):
    try:
        if value is None or str(value).strip() == '':
            cell.value = None
        else:
            cell.value = float(value)
    except ValueError:
        cell.value = value  # 숫자로 변환 불가한 경우 원래대로

# 기록 삽입
for _, row in df.iterrows():
    name = row['name']
    day = row['date'].day
    if name not in name_to_row:
        continue
    col = date_to_col.get(day)
    base_row = name_to_row[name]

    def v(x): return str(x) if pd.notna(x) and x != '' else ''

    # off 텍스트 매핑
    off_val = v(row['off'])
    if off_val == '오전반차':
      off_val = '4전반'
    elif off_val == '오후반차':
      off_val = '4후반'

    flex_val = v(row['flexOt'])

    # 병합
    combined = off_val
    if flex_val:
      combined = f"{off_val} / {flex_val}" if off_val else flex_val

    set_number_or_blank(ws[f"{col}{base_row}"], combined)
    set_number_or_blank(ws[f"{col}{base_row+1}"], v(row['ot']))
    set_number_or_blank(ws[f"{col}{base_row+2}"], v(row['holidayOt']))
    set_number_or_blank(ws[f"{col}{base_row+3}"], v(row['nightOt']))

    # note 주석 추가 
    note_val = v(row['note'])
    if note_val:
      comment = Comment(note_val, "우빈")
      ws[f"{col}{base_row}"].comment = comment

# 토/일요일 색 구분
weekend_fill = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")
for day in range(1, 32):
    try:
        dt = datetime.strptime(f"{month_str}-{day:02}", "%Y-%m-%d")
        if dt.weekday() >= 5:
            col = date_to_col[day]
            for base in name_to_row.values():
                for offset in range(4):
                    ws[f"{col}{base+offset}"].fill = weekend_fill
    except:
        continue

# 저장
out_path = os.path.join(output_dir, f"근무현황_영상의학과({target_month_code}).xlsx")
wb.save(out_path)
print(out_path)
